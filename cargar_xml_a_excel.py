from __future__ import annotations

import re
import tkinter as tk
import unicodedata
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


NS = {
    "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
    "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
    "sac": "urn:sunat:names:specification:ubl:peru:schema:xsd:SunatAggregateComponents-1",
}

TARGET_FIELDS = [
    "# DE FACTURA",
    "PROVEEDOR",
    "CARROCERIA",
    "SERIE DE LA CARROCERIA",
    "MARCA",
    "MODELO",
    "AÑO MODELO",
    "PLACA",
    "# MOTOR",
    "# DE CHASIS",
    "# DE ASIENTOS",
    "PRECIO DE VENTA",
]


def normalize_text(value: str) -> str:
    value = value or ""
    value = value.strip().upper()
    value = "".join(
        c for c in unicodedata.normalize("NFD", value) if unicodedata.category(c) != "Mn"
    )
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


HEADER_ALIASES = {
    "# DE FACTURA": ["# DE FACTURA", "N DE FACTURA", "NUMERO DE FACTURA"],
    "PROVEEDOR": ["PROVEEDOR"],
    "CARROCERIA": ["CARROCERIA"],
    "SERIE DE LA CARROCERIA": [
        "SERIE DE LA CARROCERIA",
        "CHASIS O SERIE",
        "SERIE",
    ],
    "MARCA": ["MARCA"],
    "MODELO": ["MODELO"],
    "AÑO MODELO": ["AÑO MODELO", "ANO MODELO"],
    "PLACA": ["PLACA"],
    "# MOTOR": ["# MOTOR", "N MOTOR", "NUMERO DE MOTOR"],
    "# DE CHASIS": ["# DE CHASIS", "N DE CHASIS", "NUMERO DE CHASIS"],
    "# DE ASIENTOS": ["# DE ASIENTOS", "N DE ASIENTOS", "NUMERO DE ASIENTOS"],
    "PRECIO DE VENTA": ["PRECIO DE VENTA"],
}


def text_or_empty(node: ET.Element | None) -> str:
    return "" if node is None or node.text is None else node.text.strip()


def clean_prefix(raw_value: str, prefix: str) -> str:
    value = raw_value.strip()
    if normalize_text(value).startswith(normalize_text(prefix)):
        return value[len(prefix) :].strip(" :-")
    return value


def parse_description_fields(description: str) -> dict[str, str]:
    data = {"carroceria": "", "chasis": "", "motor": "", "anio_modelo": ""}
    if not description:
        return data

    first_line = description.splitlines()[0].strip()
    if first_line:
        data["carroceria"] = first_line.split(",")[0].strip()

    chasis_match = re.search(r"CHASIS\s*:?\s*:?\s*([A-Z0-9-]+)", description, re.IGNORECASE)
    if chasis_match:
        data["chasis"] = chasis_match.group(1).strip()

    motor_match = re.search(r"MOTOR\s*:?\s*:?\s*([A-Z0-9-]+)", description, re.IGNORECASE)
    if motor_match:
        data["motor"] = motor_match.group(1).strip()

    anio_match = re.search(r"A[ÑN]O\s+MODELO\s*:?\s*:?\s*(\d{4})", description, re.IGNORECASE)
    if anio_match:
        data["anio_modelo"] = anio_match.group(1).strip()

    return data


def parse_additional_properties(root: ET.Element) -> dict[str, str]:
    properties: dict[str, str] = {}
    for prop in root.findall(".//sac:AdditionalProperty", NS):
        key = text_or_empty(prop.find("cbc:ID", NS))
        val = text_or_empty(prop.find("cbc:Value", NS))
        if key:
            properties[key] = val
    return properties


def parse_invoice(xml_path: Path) -> dict[str, str]:
    tree = ET.parse(xml_path)
    root = tree.getroot()

    additional = parse_additional_properties(root)
    invoice_number = text_or_empty(root.find("cbc:ID", NS))

    if not invoice_number:
        # Fallback usando el nombre de archivo: 01-0F303-0000379.xml -> 0F303-0000379
        m = re.match(r"^\d{2}-(.+)\.xml$", xml_path.name, flags=re.IGNORECASE)
        if m:
            invoice_number = m.group(1)

    # Normaliza IDs tipo 0F303-0000379 a F303-0000379.
    invoice_number = re.sub(r"^0([A-Z]\d{3}-\d+)$", r"\1", invoice_number.strip(), flags=re.IGNORECASE)

    proveedor = text_or_empty(
        root.find("cac:AccountingSupplierParty/cac:Party/cac:PartyName/cbc:Name", NS)
    )
    if not proveedor:
        proveedor = text_or_empty(
            root.find(
                "cac:AccountingSupplierParty/cac:Party/cac:PartyLegalEntity/cbc:RegistrationName",
                NS,
            )
        )

    description = text_or_empty(root.find("cac:InvoiceLine/cac:Item/cbc:Description", NS))
    desc = parse_description_fields(description)

    marca = clean_prefix(additional.get("02", ""), "MARCA:")
    modelo = clean_prefix(additional.get("04", ""), "MODELO:")
    motor = clean_prefix(additional.get("03", ""), "MOTOR:")
    chasis = clean_prefix(additional.get("07", ""), "NUMERO DE CHASIS:")

    if not chasis:
        chasis = desc["chasis"]
    if not motor:
        motor = desc["motor"]

    anio_modelo = desc["anio_modelo"]
    if not anio_modelo:
        anio_modelo = clean_prefix(additional.get("05", ""), "AÑO MODELO:")

    carroceria = desc["carroceria"] or "REMOLCADOR"
    total_igv = text_or_empty(root.find("cac:LegalMonetaryTotal/cbc:TaxInclusiveAmount", NS))
    if not total_igv:
        total_igv = text_or_empty(root.find("cac:LegalMonetaryTotal/cbc:PayableAmount", NS))

    return {
        "# DE FACTURA": invoice_number,
        "PROVEEDOR": proveedor,
        "CARROCERIA": carroceria,
        "SERIE DE LA CARROCERIA": chasis,
        "MARCA": marca,
        "MODELO": modelo,
        "AÑO MODELO": anio_modelo,
        "PLACA": "",
        "# MOTOR": motor,
        "# DE CHASIS": chasis,
        "# DE ASIENTOS": "",
        "PRECIO DE VENTA": total_igv,
    }


def alias_match(target_field: str, header_value: str) -> bool:
    h = normalize_text(header_value)
    aliases = [normalize_text(x) for x in HEADER_ALIASES[target_field]]
    return h in aliases


def find_header_row_and_columns(ws) -> tuple[int, dict[str, int]]:
    best_row = 0
    best_map: dict[str, int] = {}

    max_rows_to_scan = min(40, ws.max_row if ws.max_row > 0 else 40)
    max_cols_to_scan = min(120, ws.max_column if ws.max_column > 0 else 120)

    for row in range(1, max_rows_to_scan + 1):
        current_map: dict[str, int] = {}
        for col in range(1, max_cols_to_scan + 1):
            raw = ws.cell(row=row, column=col).value
            if raw is None:
                continue
            cell_text = str(raw)
            for field in TARGET_FIELDS:
                if field not in current_map and alias_match(field, cell_text):
                    current_map[field] = col
        if len(current_map) > len(best_map):
            best_map = current_map
            best_row = row

    if len(best_map) < 6:
        raise ValueError(
            "No se detectaron suficientes encabezados. Verifica que la hoja tenga los títulos esperados."
        )

    return best_row, best_map


def find_best_sheet_and_columns(wb):
    best_ws = None
    best_header_row = 0
    best_map: dict[str, int] = {}

    for ws in wb.worksheets:
        try:
            header_row, col_map = find_header_row_and_columns(ws)
            if len(col_map) > len(best_map):
                best_ws = ws
                best_header_row = header_row
                best_map = col_map
        except ValueError:
            continue

    if best_ws is None:
        raise ValueError(
            "No se detectaron encabezados validos en ninguna hoja del Excel."
        )

    return best_ws, best_header_row, best_map


def find_next_data_row(ws, header_row: int, mapped_columns: dict[str, int]) -> int:
    # Inserta siempre al final del bloque de datos ya registrado.
    factura_col = mapped_columns.get("# DE FACTURA")

    if factura_col is not None:
        last_factura_row = header_row
        for row in range(header_row + 1, ws.max_row + 1):
            value = ws.cell(row=row, column=factura_col).value
            if value not in (None, ""):
                last_factura_row = row
        return last_factura_row + 1

    candidate_cols = list(mapped_columns.values())
    last_data_row = header_row
    for row in range(header_row + 1, ws.max_row + 1):
        if any(ws.cell(row=row, column=c).value not in (None, "") for c in candidate_cols):
            last_data_row = row
    return last_data_row + 1


class LoaderApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Cargar XML a tabla Excel")
        self.root.geometry("860x480")

        self.excel_path: Path | None = None
        self.xml_paths: list[Path] = []

        self._build_ui()

    def _build_ui(self) -> None:
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text="Carga XML y escribe directo en tu tabla", font=("Segoe UI", 12, "bold")).pack(
            anchor="w"
        )

        btn_row = ttk.Frame(main)
        btn_row.pack(fill="x", pady=(10, 8))

        ttk.Button(btn_row, text="1) Seleccionar Excel", command=self.select_excel).pack(side="left")
        ttk.Button(btn_row, text="2) Seleccionar XML", command=self.select_xml).pack(side="left", padx=8)
        ttk.Button(btn_row, text="3) Cargar a tabla", command=self.load_to_excel).pack(side="left")

        self.lbl_excel = ttk.Label(main, text="Excel: (no seleccionado)")
        self.lbl_excel.pack(anchor="w", pady=(4, 2))

        self.lbl_xml = ttk.Label(main, text="XML: 0 seleccionados")
        self.lbl_xml.pack(anchor="w", pady=(0, 8))

        self.log = tk.Text(main, height=18, wrap="word")
        self.log.pack(fill="both", expand=True)
        self._append_log("Listo. Selecciona primero el Excel y luego tus XML.")

    def _append_log(self, line: str) -> None:
        self.log.insert("end", line + "\n")
        self.log.see("end")

    def select_excel(self) -> None:
        path = filedialog.askopenfilename(
            title="Selecciona tu archivo Excel",
            filetypes=[("Excel", "*.xlsx *.xlsm")],
            initialdir=str(Path(__file__).resolve().parent),
        )
        if not path:
            return
        self.excel_path = Path(path)
        self.lbl_excel.config(text=f"Excel: {self.excel_path.name}")
        self._append_log(f"Excel seleccionado: {self.excel_path}")

    def select_xml(self) -> None:
        paths = filedialog.askopenfilenames(
            title="Selecciona XML de facturas",
            filetypes=[("XML", "*.xml")],
            initialdir=str(Path(__file__).resolve().parent),
        )
        if not paths:
            return
        self.xml_paths = [Path(p) for p in paths]
        self.lbl_xml.config(text=f"XML: {len(self.xml_paths)} seleccionados")
        self._append_log(f"XML seleccionados: {len(self.xml_paths)}")

    def load_to_excel(self) -> None:
        if self.excel_path is None:
            messagebox.showwarning("Falta Excel", "Primero selecciona tu archivo Excel.")
            return
        if not self.xml_paths:
            messagebox.showwarning("Faltan XML", "Primero selecciona uno o varios XML.")
            return

        try:
            keep_vba = self.excel_path.suffix.lower() == ".xlsm"
            wb = load_workbook(self.excel_path, keep_vba=keep_vba)
            ws, header_row, col_map = find_best_sheet_and_columns(wb)
            next_row = find_next_data_row(ws, header_row, col_map)

            self._append_log(f"Hoja detectada: {ws.title}")
            self._append_log(f"Fila inicial de carga: {next_row}")

            loaded = 0
            failed = 0
            for xml_path in self.xml_paths:
                try:
                    data = parse_invoice(xml_path)
                    if not data.get("# DE FACTURA", "").strip():
                        raise ValueError("No se encontro # DE FACTURA en el XML")
                    for field, value in data.items():
                        col = col_map.get(field)
                        if col is not None:
                            ws.cell(row=next_row, column=col, value=value)
                    loaded += 1
                    self._append_log(f"OK  {xml_path.name} -> fila {next_row}")
                    next_row += 1
                except Exception as ex:
                    failed += 1
                    self._append_log(f"ERR {xml_path.name}: {ex}")

            try:
                wb.save(self.excel_path)
                self._append_log(f"Guardado OK en: {self.excel_path}")
                messagebox.showinfo(
                    "Proceso terminado",
                    f"Cargados: {loaded}\nCon error: {failed}\nArchivo: {self.excel_path.name}",
                )
            except PermissionError:
                self._append_log("No se pudo guardar: el Excel esta abierto o bloqueado.")

                save_as = filedialog.asksaveasfilename(
                    title="El archivo esta en uso. Guardar una copia como...",
                    defaultextension=self.excel_path.suffix,
                    filetypes=[("Excel", "*.xlsx *.xlsm")],
                    initialdir=str(self.excel_path.parent),
                    initialfile=f"{self.excel_path.stem}_actualizado{self.excel_path.suffix}",
                )

                if save_as:
                    wb.save(save_as)
                    self._append_log(f"Guardado como copia: {save_as}")
                    messagebox.showinfo(
                        "Guardado como copia",
                        "El archivo original estaba abierto.\n"
                        f"Se guardo una copia con los cambios en:\n{save_as}",
                    )
                else:
                    messagebox.showwarning(
                        "No guardado",
                        "Se cargaron los XML en memoria, pero no se guardo porque el archivo estaba en uso.\n"
                        "Cierra el Excel y vuelve a ejecutar 'Cargar a tabla'.",
                    )
            finally:
                wb.close()
        except Exception as ex:
            messagebox.showerror("Error", str(ex))


def main() -> None:
    root = tk.Tk()
    app = LoaderApp(root)
    _ = app
    root.mainloop()


if __name__ == "__main__":
    main()